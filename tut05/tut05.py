import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
path2 = './output'

isfile2 = os.path.isdir(path2)
if(isfile2 == False):
    os.mkdir(path2)

for f in os.listdir(path2):
    os.remove(os.path.join(path2, f))

subno_to_sunbname = {}
subno_to_ltp = {}
roll_to_name = {}
grade = {"AA": 10, "AB": 9, "BB": 8, "BC": 7,
         "CC": 6, "CD": 5, "DD": 4, "F": 0, "I": 0}
st = {"AA*": "AA", "AB*": "AB", "BB*": "BB", "BC*": "BC", "CC*": "CC", "CD*": "CC", "DD*": "DD", "F*": "F", "I*": "I",
      "AA": "AA", "AB": "AB", "BB": "BB", "BC": "BC", "CC": "CC", "CD": "CC", "DD": "DD", "F": "F", "I": "I"}
with open("subjects_master.csv", 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        dct = dict(row)
        subno_to_sunbname[dct['subno']] = dct['subname']
        subno_to_ltp[dct['subno']] = dct['ltp']
with open("names-roll.csv", 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        dct = dict(row)
        roll_to_name[dct['Roll']] = dct['Name']


def generate_marksheet():

    with open("grades.csv", 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            dct = dict(row)
            input_filename = dct['Roll'] + ".xlsx"

            if os.path.isfile("./output/"+input_filename) == False:

                wb = Workbook()
                wb.save("./output/"+input_filename)

            book = load_workbook("./output/"+input_filename)

            if 'Sem'+dct['Sem'] not in book.sheetnames:
                book.create_sheet('Sem'+dct['Sem'])
                Sheet1 = book['Sem'+dct['Sem']]
                Sheet1.append(["Sl No.", "Subject No.",
                               "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade"])
            Sheet1 = book['Sem'+dct['Sem']]
            row_count = Sheet1.max_row
            Sheet1.append([row_count, dct['SubCode'],
                           subno_to_sunbname[dct['SubCode']], subno_to_ltp[dct['SubCode']], dct['Credit'], dct['Sub_Type'], dct['Grade']])
            book.save("./output/"+input_filename)

    return


def overall():
    for f in os.listdir(path2):
        book = load_workbook("./output/"+f)
        if "Sheet" in book.sheetnames:
            ss = book['Sheet']
            ss.title = 'Overall'
        ss = book['Overall']
        ss.append(["Roll No.", f[:-5]])
        ss.append(["Name of Student", roll_to_name[f[:-5]]])
        ss.append(["Discipline", f[4:6]])
        ss["A4"] = "Semester No."
        ss["A5"] = "Semester wise Credit Taken"
        ss["A6"] = "SPI"
        ss["A7"] = "Total Credits Taken"
        ss["A8"] = "CPI"
        t = 0
        tt = 0
        x = 1
        for nn in book.sheetnames:
            if nn == 'Overall':
                continue
            s = book[nn]
            ss.cell(row=4, column=x+1).value = nn[3:]
            sm = 0
            spi = 0
            for a in range(2, s.max_row + 1):
                sm += int(s.cell(row=a, column=5).value)
                g = st[s.cell(row=a, column=7).value.strip()]
                spi += (int(s.cell(row=a, column=5).value))*grade[g]
            ss.cell(row=5, column=x+1).value = sm
            spi = spi/sm
            ss.cell(row=6, column=x+1).value = round(spi, 2)
            t += sm
            tt += spi*sm
            ss.cell(row=7, column=x+1).value = t
            ss.cell(row=8, column=x+1).value = round(tt/t, 2)
            x += 1

        book.save("./output/"+f)


generate_marksheet()
overall()
