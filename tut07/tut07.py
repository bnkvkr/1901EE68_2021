from openpyxl import load_workbook
import openpyxl
import os
import csv
import re
roll_inf = {}
sub_ltp = {}
fd = {}


def final(result):

    with open("course_registered_by_all_students.csv") as file:
        reader = csv.DictReader(file)

        try:
            wb = load_workbook('course_feedback_remaining.xlsx')
        except:
            fpath = 'course_feedback_remaining.xlsx'
            wb = openpyxl.Workbook()
            wb.save(fpath)
            ss = wb['Sheet']
            ss.title = 'Sheet1'
            ww = wb['Sheet1']
            ww.append(["rollno.", "register_sem", "schedule_sem",
                       "subno", "Name", "email", "aemail", "contact"])
        fpath = 'course_feedback_remaining.xlsx'

        for row in reader:

            dct = dict(row)

            cc = []
            if dct['rollno'] in fd.keys():
                cc = fd[dct['rollno']]
            aa = sub_ltp[dct['subno']]

            for i in range(1, 4):
                if(aa[i-1] != '0'):
                    j = i

                    bb = [dct['subno'], str(j)]
                    f = 0
                    for dd in cc:
                        if(dd == bb):
                            f = 1
                            break
                    if f == 0:
                        if dct['rollno'] not in roll_inf.keys():
                            ee = [dct['rollno'], dct['register_sem'],
                                  dct['schedule_sem'], dct['subno']]
                            for x in range(0, 4):
                                ee.append('NA_IN_STUDENTINFO')
                            result.append(ee)
                        else:
                            ee = [dct['rollno'], dct['register_sem'],
                                  dct['schedule_sem'], dct['subno']]
                            ee.extend(roll_inf[dct['rollno']])
                            result.append(ee)
                        break

        ww = wb['Sheet1']
        for rr in result:
            ww.append(rr)
        wb.save(fpath)


def save_roll_info():

    with open("studentinfo.csv") as file:
        reader = csv.DictReader(file)
        for row in reader:
            dct = dict(row)
            temp = []
            temp.append(dct['Name'])
            temp.append(dct['email'])
            temp.append(dct['aemail'])
            temp.append(dct['contact'])
            roll_inf[dct['Roll No']] = temp


def save_sub_ltp():
    with open("course_master_dont_open_in_excel.csv") as file:
        reader = csv.DictReader(file)
        for row in reader:
            dct = dict(row)
            ss = dct['ltp']
            aa = ss.split('-')
            sub_ltp[dct['subno']] = aa


def feeding():
    with open("course_feedback_submitted_by_students.csv") as file:
        reader = csv.DictReader(file)
        for row in reader:
            dct = dict(row)
            temp = []
            temp.append(dct['course_code'])
            temp.append(dct['feedback_type'])
            if dct['stud_roll'] not in fd.keys():
                fd[dct['stud_roll']] = []
            fd[dct['stud_roll']].append(temp)


def feedback_not_submitted():
    output_file_name = "course_feedback_remaining.xlsx"
    if os.path.exists(output_file_name):
        os.remove(output_file_name)
    save_roll_info()
    result = []
    save_sub_ltp()
    feeding()
    final(result)


feedback_not_submitted()
