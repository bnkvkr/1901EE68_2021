from os.path import isfile, join
from flask import Flask, render_template, request, jsonify, flash,  send_file, flash, redirect, url_for
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
import openpyxl
from fpdf import FPDF
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import pytz
from os import listdir
import acc
import acc1
from flask import jsonify
app = Flask(__name__)
seal = ''
sign = ''
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'

path2 = './sign'
isfile2 = os.path.isdir(path2)
if(isfile2 == False):
    os.mkdir(path2)

path1 = './seal'
isfile2 = os.path.isdir(path1)
if(isfile2 == False):
    os.mkdir(path1)


def pree():
    path2 = './transcriptsIITP'

    isfile2 = os.path.isdir(path2)
    if(isfile2 == False):
        os.mkdir(path2)

    for f in os.listdir(path2):
        os.remove(os.path.join(path2, f))


@app.route("/")
def hello_world():
    return render_template('index.html', a=[], d="style ='display:none'", ln=0)


@app.route("/upload", methods=['GET', 'POST'])
def upload1():

    if request.method == 'POST':

        try:

            f = request.files['file']
            print(f, f.filename)
           # print('hello')
            f.save(f.filename)
        except:
            print("You have not selected any file 😲")
            return render_template('index.html')
        flash(f"{f.filename} uploaded successfully 👍")

    return render_template('index.html', a=[], d="style ='display:none'", ln=0)


@app.route("/upload2", methods=['GET', 'POST'])
def uploadsign():
    path2 = './sign'
    isfile2 = os.path.isdir(path2)
    if(isfile2 == False):
        os.mkdir(path2)
    for f in os.listdir(path2):
        os.remove(os.path.join(path2, f))

    if request.method == 'POST':

        try:

            f = request.files['file']
            f.save(f'./sign/{f.filename}')
            f.save(f.filename)
        except:
            print("You have not selected any file 😲")
            return render_template('index.html')
        flash(f"{f.filename} uploaded successfully 👍")

    return render_template('index.html', a=[], d="style ='display:none'", ln=0)


@app.route("/upload3", methods=['GET', 'POST'])
def uploadseal():

    path1 = './seal'
    isfile2 = os.path.isdir(path1)
    if(isfile2 == False):
        os.mkdir(path1)
    for f in os.listdir(path1):
        os.remove(os.path.join(path1, f))

    if request.method == 'POST':

        try:
            f = request.files['file']
            #print(f, f.filename)
            print(sign)
            seal = f.filename
            print(seal)
            f.save(f'./seal/{f.filename}')
        except:
            print("You have not selected any file 😲")
            return render_template('index.html')
        flash(f"{f.filename} uploaded successfully 👍")

    return render_template('index.html', a=[], d="style ='display:none'", ln=0)


@app.route("/generate-range", methods=['GET', 'POST'])
def generate():
    pree()
    if request.method == 'POST':
        roll_list = []
        first = request.form['First'].upper()
        second = request.form['Second'].upper()
        # if(first.strip()=="" or second.strip()==""):
        #     flash("Enter a valid Range")

        f = first[::-1]
        s = second[::-1]
        i = 0
        l = ""
        r = ""
        while 1:
            if(f[i] >= '0' and f[i] <= '9'):
                l = f[i]+l
            else:
                break
            i += 1

        x = first[:len(first)-i]
        i = 0
        while 1:
            if(s[i] >= '0' and s[i] <= '9'):
                r = s[i]+r
            else:
                break
            i += 1
        l = int(l)
        r = int(r)
        for i in range(l, r+1):
            if(len(str(i)) == 1):
                roll_list.append((x+'0'+str(i)))
            else:
                roll_list.append(x+str(i))

        if first[2] + first[3] == "01":
            a = acc.btech(roll_list)
            print(a)
            # all_roll(roll_list)
        else:
            a = acc1.all_roll(roll_list)
            print(a)
    # flash("Roll No Generated successfully")
    return jsonify(a=a, ln=len(a))


@app.route("/generate", methods=['GET', 'POST'])
def generateall():
    pree()
    btech_roll = []
    other_roll = []
    with open("names-roll.csv", 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            dct = dict(row)
            first = dct['Roll']
            if first[2] + first[3] == "01":
                btech_roll.append(first.upper())
            else:
                other_roll.append(first.upper())
    a = acc.btech(btech_roll)
    b = acc1.all_roll(other_roll)
    print(a)
    print(b)
    path1 = './sign'

    for f in os.listdir(path1):
        os.remove(os.path.join(path1, f))
    path1 = './seal'

    for f in os.listdir(path1):
        os.remove(os.path.join(path1, f))

    return 'Sucesss', 200


if __name__ == "__main__":
    app.run(debug=True)
