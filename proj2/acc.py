from fpdf import FPDF
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import pytz
from os.path import isfile, join
from flask import Flask, render_template, request, jsonify, flash,  send_file, flash, redirect, url_for
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
import openpyxl
from fpdf import FPDF
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import pytz
from os import listdir
subno_to_sunbname = {}
subno_to_ltp = {}
roll_to_name = {}
rolls = []
branch_to_fs = {"CS": "Computer Science and Engineering", "EE": "Electrical Engineering", "ME": "Mechanical Engineering",
                "CE": "Civil and Environmental Engineering", "CBE": "Chemical and Biochemical Engineering", "MME": "Metallurgical and Materials Engineering"}
stream = {"01": "Bachelor of Technology", "11": "Master of Technology",
          "12": "Master of Science", "21": "Doctor of Philosophy"}
grade = {"AA": 10, "AB": 9, "BB": 8, "BC": 7,
         "CC": 6, "CD": 5, "DD": 4, "F": 0, "I": 0}
st = {"AA*": "AA", "AB*": "AB", "BB*": "BB", "BC*": "BC", "CC*": "CC", "CD*": "CC", "DD*": "DD", "F*": "F", "I*": "I",
      "AA": "AA", "AB": "AB", "BB": "BB", "BC": "BC", "CC": "CC", "CD": "CC", "DD": "DD", "F": "F", "I": "I"}


def btech(roll_list):

    # time = x.strftime("%d %b %Y, %H:%M")
    # print(x.strftime("%d %b %Y, %H:%M"))
    non_exist = []
    non_exist.clear()
    subno_to_sunbname = {}
    subno_to_ltp = {}
    roll_to_name = {}
    branch_to_fs = {"CS": "Computer Science and Engineering", "EE": "Electrical Engineering", "ME": "Mechanical Engineering",
                    "CE": "Civil and Environmental Engineering", "CBE": "Chemical and Biochemical Engineering", "MME": "Metallurgical and Materials Engineering"}
    stream = {"01": "Bachelor of Technology", "11": "Master of Technology",
              "12": "Master of Science", "21": "Doctor of Philosophy"}
    rolls = []

    grade = {"AA": 10, "AB": 9, "BB": 8, "BC": 7,
             "CC": 6, "CD": 5, "DD": 4, "F": 0, "I": 0}
    st = {"AA*": "AA", "AB*": "AB", "BB*": "BB", "BC*": "BC", "CC*": "CC", "CD*": "CC", "DD*": "DD", "F*": "F", "I*": "I",
          "AA": "AA", "AB": "AB", "BB": "BB", "BC": "BC", "CC": "CC", "CD": "CC", "DD": "DD", "F": "F", "I": "I"}
    with open("subjects_master.csv", 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            dct = dict(row)
            subno_to_sunbname[dct['subno']] = dct['subname']
            subno_to_ltp[dct['subno']] = dct['ltp']

    with open("names-roll.csv", 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            dct = dict(row)
            roll_to_name[dct['Roll'].upper()] = dct['Name']
           # rolls.append(dct['Roll'].upper())

    for roll in roll_list:
        if roll in roll_to_name.keys():
            rolls.append(roll)
        else:
            non_exist.append(roll)

    # rolls=['0401CS01','0401CS02','0401EE19']
    for roll in rolls:

        with open("grades.csv", 'r') as file:
            reader = csv.DictReader(file)
            sheet = defaultdict(list)
            pdf = FPDF('L', 'mm', 'A3')

            pdf.add_page()

            pdf.rect(10, 10, 400, 275, 'D')  # outer box
            pdf.rect(10, 10, 60, 30, 'D')
            pdf.rect(10, 10, 340, 30, 'D')
            pdf.rect(10, 10, 400, 30, 'D')

            pdf.rect(10, 10, 400, 105, 'D')

            pdf.rect(115, 42, 185, 13, 'D')

            pdf.rect(10, 10, 400, 170, 'D')
            pdf.rect(10, 10, 400, 228, 'D')
            pdf.image("iitp_1.png", 28, 11, 25)
            pdf.set_font('helvetica', 'BU', 8)
            pdf.cell(60, 55, 'INTERIM TRANSCRIPT', align='C')

            pdf.image("iitp_1.png", 367, 11, 25)
            pdf.cell(195)
            pdf.cell(228, 55, 'INTERIM TRANSCRIPT', align='C')

            pdf.add_font('FreeSerif', '', 'FreeSerif.ttf', uni=True)
            pdf.set_font('FreeSerif', '', 20)
            pdf.set_xy(90, 10)
            # pdf.set_y(15)
            pdf.image("hindi.jpg", 132, 10.5, 160, 13.5)
            #pdf.cell(10, 11, 'भारतीय प्रौद्योगिकी संस्थान पटना',ln=1)
            pdf.set_font('FreeSerif', '', 25)
            pdf.set_xy(125, 2)
            pdf.cell(168, 47, 'Indian Institue of Technology Patna', align='C')
            pdf.set_xy(125, 10)
            pdf.cell(168, 48, 'Transcript', align='C')
            pdf.set_xy(60, 25)
            pdf.set_font('helvetica', 'B', 10)
            pdf.cell(147, 41, 'Roll No:', align='C')
            pdf.rect(144, 44, 24, 4, 'D')

            pdf.cell(-40, 41, 'Name:', align='C')
            pdf.rect(195, 44, 40, 4, 'D')
            pdf.cell(185, 41, 'Year of Admission:', align='C')
            pdf.rect(280, 44, 12, 4, 'D')
            pdf.set_xy(115, 23)
            pdf.cell(
                168, 57, f'Programmme: {stream[roll[2]+roll[3]]}      Course: {branch_to_fs[roll[4]+roll[5]]} ', align='C')
            # pdf.set_xy(115,43)
            pdf.set_x(105)
            pdf.cell(100, 47, f'{roll}', align='C')
            pdf.cell(15, 47, f'{roll_to_name[roll]}', align='C')
            pdf.cell(130, 47, f'{"20"+roll[0]+roll[1]}', align='C')
            pdf.set_font('helvetica', '', 5)

            for row in reader:
                dct = dict(row)
                # print(dct)

                if(dct['Roll'].upper() == roll):
                    sheet[dct['Sem']].append([dct['SubCode'], subno_to_sunbname[dct['SubCode']],
                                             subno_to_ltp[dct['SubCode']], dct['Credit'], dct['Grade']])

            # print(sheet)

            j = 0

            c = 0
            v = 60
            cnt = 0

            cpi = 0
            it = 0
            tot = 0

            # print(sheet)
            for key in sheet:
                cnt += 1
                req = ''
                data = []
                cre = 0
                spi = 0

                data.append(
                    ["Sub Code",  "Subject Name", "L-T-P", "CRD", "GRD"])

                spi_c = 0
                cleared = 0
                for a in sheet[key]:
                    cre += int(a[3])
                    spi_c += int(a[3])*grade[st[a[4].strip()]]
                    if(grade[st[a[4].strip()]] >= 5):
                        cleared += int(a[3])
                    data.append(a)
                tot += cre
                spi = spi_c/cre
                cpi += spi*cre
                line_height2 = pdf.font_size * 2.5
                line_height1 = pdf.font_size * 2.5

                col_width = pdf.epw / 23
                col_width1 = pdf.epw / 25
                col_width2 = pdf.epw/7

                pdf.set_y(v-2)
                flag = 1
                flag2 = 1
                x = 0
                pdf.ln(3)
                if(c == 0):
                    x = 18

                elif c == 1:
                    x = 150

                else:
                    x = 282

                pdf.set_x(x-3)
                # pdf.set_y(v-5)
                pdf.set_font('helvetica', 'BU', 8)
                pdf.multi_cell(col_width1+10, line_height1,
                               f'Semester {key}', border=0, ln=3, max_line_height=pdf.font_size, align='C')
                pdf.ln(4)
                pdf.set_font('helvetica', '', 5)
                for row in data:
                    if(c == 0):
                        pdf.cell(10)
                    elif c == 1:
                        pdf.cell(5*28.5)
                    else:
                        pdf.cell(10*27.5)
                    if flag:
                        pdf.set_font(style="B")
                        flag = 0
                    else:
                        pdf.set_font(style="")

                    pdf.multi_cell(
                        col_width1, line_height1, row[0], border=1, ln=3, max_line_height=pdf.font_size, align='C')
                    pdf.multi_cell(
                        col_width2, line_height2, row[1], border=1, ln=3, max_line_height=pdf.font_size, align='C')
                    pdf.multi_cell(
                        col_width-5, line_height1, row[2], border=1, ln=3, max_line_height=pdf.font_size, align='C')
                    pdf.multi_cell(
                        col_width-5, line_height1, row[3], border=1, ln=3, max_line_height=pdf.font_size, align='C')
                    pdf.multi_cell(
                        col_width-5, line_height2, row[4], border=1, ln=3, max_line_height=pdf.font_size, align='C')
                    pdf.ln(line_height1)
                j += 5

                x = 0
                pdf.ln(3)
                if(c == 0):
                    x = 17

                elif c == 1:
                    x = 149

                else:
                    x = 281.5

                pdf.set_x(x+3)

                req = f'Credits Taken: {cre}    Cleared: {cleared}    SPI: {round(spi,2)}    CPI: {round(cpi/tot,2)}'
                pdf.multi_cell(col_width2+5, line_height2, req, border=1,
                               ln=3, max_line_height=pdf.font_size, align='C')
                c += 1
                it += 1
                if(c == 3):
                    c = 0
                    j = 0
                    if(v == 60):
                        v = 122
                    else:
                        v = 185
            IST = pytz.timezone('Asia/Kolkata')
            xx = datetime.now(IST)
            # print(x.strftime("%d %b %Y, %H:%M"))
            # pdf.set_y(216)
            # print(len(seal),len(sign),seal,sign)

            mypath = './seal'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

            # print(onlyfiles)

            if(len(onlyfiles) > 0):
                pdf.image(f'./seal/{onlyfiles[0]}', 198, 250.5, 24, 24.5)
            pdf.set_xy(20, 227)
            # pdf.ln(2)
            # pdf.SetAutoPageBreak(False, 0)
            pdf.set_font('helvetica', '', 17)
            pdf.cell(10, 50, 'Date Generated: ', align='')
            pdf.set_font('helvetica', 'BU', 20)
            pdf.set_x(68)
            pdf.cell(50, 50, f'{xx.strftime("%d %b %Y, %H:%M")}', align='')
            #pdf.cell(40,50, 'Assistant Registrar(Academic): ', align='')

            mypath2 = './sign'
            onlyfiles2 = [f for f in listdir(
                mypath2) if isfile(join(mypath2, f))]
            if(len(onlyfiles2) > 0):
                pdf.image(f'./sign/{onlyfiles2[0]}', 355, 245, 24, 24.5)
            # else:
            #     pdf.image("rovi.jpeg", 355, 245, 24, 24.5)
            pdf.set_font('helvetica', '', 17)
            pdf.set_x(307)

            pdf.image("Assign.jpg", 340, 268, 64, 8)
            pdf.output(f'./transcriptsIITP/{roll}.pdf')
    return non_exist
