import openpyxl
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
path2 = './output_by_subject'
isfile2 = os.path.isdir(path2)
if(isfile2 == False):
    os.mkdir(path2)
else:
    for f in os.listdir(path2):
        os.remove(os.path.join(path2, f))

path1 = './output_individual_roll'
isFile = os.path.isdir(path1)
if(isFile == False):
    os.mkdir(path1)
else:
    for f in os.listdir(path1):
        os.remove(os.path.join(path1, f))


def output_individual_roll():
    with open('regtable_old.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            data = [row[0],  row[1], row[3], row[8]]
            input_filename = row[0] + ".xlsx"
            if os.path.isfile("./output_individual_roll/"+input_filename):
                book = openpyxl.load_workbook(
                    "./output_individual_roll/"+input_filename)
                Sheet1 = book.active
                Sheet1.append(data)
                book.save("./output_individual_roll/"+input_filename)
            else:
                wb = Workbook()
                Sheet1 = wb.active
                Sheet1.append(["rollno", "register_sem", "subno", "sub_type"])
                Sheet1.append(data)
                wb.save("./output_individual_roll/"+input_filename)

    return


def output_by_subject():
    with open('regtable_old.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            data = [row[0],  row[1], row[3], row[8]]
            input_filename = row[3] + ".xlsx"
            if os.path.isfile("./output_by_subject/"+input_filename):
                book = openpyxl.load_workbook(
                    "./output_by_subject/"+input_filename)
                Sheet1 = book.active
                Sheet1.append(data)
                book.save("./output_by_subject/"+input_filename)
            else:
                wb = Workbook()
                Sheet1 = wb.active
                Sheet1.append(["rollno", "register_sem", "subno", "sub_type"])
                Sheet1.append(data)
                wb.save("./output_by_subject/"+input_filename)

    return


output_individual_roll()
output_by_subject()
