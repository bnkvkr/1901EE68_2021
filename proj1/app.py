from flask import Flask, render_template, request, jsonify,  send_file, flash, redirect, url_for
import csv
from openpyxl import Workbook
import os
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
import smtplib
from email.message import EmailMessage
import shutil
app = Flask(__name__)
roll_to_list = {}
roll_to_email = {}
roll_to_all = {}
list_of_title = []
list_of_roll_to_score = {}
len_of_consise = 0
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'


def anslist():
    roll_to_list.clear()
    roll_to_all.clear()
    with open("responses.csv", 'r') as file:

        reader1 = csv.reader(file)
        flag = 0
        for r in reader1:
            arr = []
            for i in range(7, len(r)):
                arr.append(r[i])

            roll_to_list[r[6].upper()] = arr
            if(flag == 1):
                roll_to_all[r[6].upper()] = r
            flag = 1


def pre_3():
    roll_to_email.clear()
    with open("responses.csv", 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            dct = dict(row)
            roll_to_email[dct['Roll Number'].upper()] = [
                dct['Email address'], dct['IITP webmail']]


def pre_2():
    roll_to_email.clear()
    with open("responses.csv", 'r') as file:
        reader = csv.DictReader(file)
        anslist()

        flag = 1
        for row in reader:
            dct = dict(row)
            if(flag):
                flag = 0
                for key in dct:
                    list_of_title.append(key)
            # input_filename = dct['Roll Number'].upper() + ".xlsx"
            roll_to_email[dct['Roll Number'].upper()] = [
                dct['Email address'], dct['IITP webmail']]


def generate_concise():
    path2 = './marksheet'

    isfile2 = os.path.isdir(path2)
    if(isfile2 == False):
        os.mkdir(path2)
    pre_2()

    if os.path.isfile("./marksheet/concise_marksheet.xlsx") == True:

        os.remove("./marksheet/concise_marksheet.xlsx")

    wb = Workbook()
    wb.save("./marksheet/concise_marksheet.xlsx")

    book = load_workbook("./marksheet/concise_marksheet.xlsx")
    # print(list_of_title)
    # print(roll_to_all)
    # print(roll_to_list)
    # print(list_of_roll_to_score)
    for worksheets in book.sheetnames:
        ss_sheet = book[worksheets]
        ss_sheet.title = 'concise'
        Sheet1 = book['concise']
        print(f"{len(list_of_title)} in gen concise")
        list_of_title[2] = 'Google_Score'
        list_of_title.insert(6, 'Score_After_Negative')
        list_of_title.pop(-1)
        for i in range(0, len(roll_to_list['ANSWER'])):
            list_of_title.append('Unnamed:')
        list_of_title.append('statusAns')
        Sheet1.append(list_of_title)
        for key in roll_to_all:

            a = roll_to_all[key]
            score = list_of_roll_to_score[key]
            a.insert(6, score[0])
            a.append('['+str(score[1])+','+str(score[2])+','+str(score[3])+']')
            a[2]=score[4]
            Sheet1.append(a)
            global len_of_consise
            len_of_consise = len(a)
        book.save("./marksheet/concise_marksheet.xlsx")
        break
    list_of_title.clear()


@app.route("/concise", methods=['POST'])
def concise():
    # print(f"{len(list_of_title)} in concise")
    # generate_concise()
    list_of_title.clear()
    if request.form['correct'] == "" or request.form['wrong'] == "":
        flash("Please enter valid input")
        return render_template('index.html')
    pm = float(request.form['correct'])
    nm = float(request.form['wrong'])
    consxx(pm, nm)

    print("Concise generated Successfully")
    flash("Concise Generated successfully")
    return render_template('index.html', x="")


@app.route("/download", methods=['POST'])
def download():

    if os.path.exists("./marksheet.zip"):
        os.remove("./marksheet.zip")

    shutil.make_archive("marksheet", 'zip', "./marksheet")

    path = "./marksheet.zip"

    return send_file(path, as_attachment=True)


@app.route("/email", methods=['POST'])
def mail():
    roll_to_email.clear()
    pre_3()
    sender_email = "rohanvishal688@gmail.com"

    sender_pass = "vishalku1@"
    receivers_email = "rohan_1901ee48@iitp.ac.in"
    sub = ("Report Card")

    cnttt = 1
    for key in roll_to_email:
        print(cnttt)

        cnttt += 1
        a = roll_to_email[key]
        files = f"./marksheet/{key}.xlsx"

        msg = EmailMessage()
        with open(files, 'rb') as f:

            file_data = f.read()
            msg['From'] = sender_email
            msg['To'] = a[0]
            msg['Cc'] = a[1]
            msg['Subject'] = sub
            msg.set_content("Hello there!! I have something for you.")
            msg.add_attachment(file_data, maintype='application',
                               subtype='octet-stream', filename=key+".xlsx")

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:

            smtp.login(sender_email, sender_pass)

            smtp.send_message(msg)

    flash("Email Sent Successfully")

    return render_template('index.html', x="")


name_roll = {}


def pre():

    path2 = './marksheet'

    isfile2 = os.path.isdir(path2)
    if(isfile2 == False):
        os.mkdir(path2)

    for f in os.listdir(path2):
        os.remove(os.path.join(path2, f))


def concise2():
    if 'ANSWER' not in roll_to_list.keys():
        return 0

    with open("master_roll.csv", 'r') as file:
        reader = csv.DictReader(file)

        f = []
        for i in range(len_of_consise):
            f.append('Absent')
        # print(len_of_consise)
        for row in reader:
            dct = dict(row)
            # print(roll_to_all)
            if(dct['roll'].upper() not in roll_to_all.keys()):

                book = load_workbook("./marksheet/concise_marksheet.xlsx")
                for worksheets in book.sheetnames:
                    ss_sheet = book[worksheets]
                    ss_sheet.title = 'concise'
                    Sheet1 = book['concise']
                    f[3] = dct['name']
                    f[7] = dct['roll'].upper()
                    print(f)
                    Sheet1.append(f)
                    book.save("./marksheet/concise_marksheet.xlsx")
                    break


def consxx(pm, nm):
    list_of_roll_to_score.clear()
    with open("responses.csv", 'r') as file:
        reader = csv.DictReader(file)

        anslist()
        ans = []
        for row in reader:
            dct = dict(row)

            if 'ANSWER' in roll_to_list.keys():
                ans = roll_to_list["ANSWER"]
            else:
                flash("No Roll Number with 'ANSWER' is present, Cannot Process!")
                return render_template('index.html')
            student = roll_to_list[dct['Roll Number'].upper()]
            cnt = 0
            na = 0

            for i in range(0, len(ans)):
                cnt += (student[i] == ans[i])
                if student[i] == "":
                    na += 1
            wng = len(ans) - cnt - na
            score = [str(pm*cnt + nm*wng) + "/" +
                        str(pm*len(ans)), cnt, wng, na,str(pm*cnt) + "/" + str(pm*len(ans))]
            list_of_roll_to_score[dct['Roll Number'].upper()] = score
    generate_concise()
    concise2()


def generate(pm, nm):

    with open("responses.csv", 'r') as file:
        reader = csv.DictReader(file)
        anslist()
        ans = []
        if 'ANSWER' in roll_to_list.keys():
            ans = roll_to_list["ANSWER"]
        else:
            flash("No Roll Number with 'ANSWER' is present, Cannot Process!")
            return render_template('index.html')

        flag = 1
        for row in reader:
            dct = dict(row)
            if(flag):
                flag = 0
                for key in dct:
                    list_of_title.append(key)
            input_filename = dct['Roll Number'].upper() + ".xlsx"
            roll_to_email[dct['Roll Number'].upper()] = [
                dct['Email address'], dct['IITP webmail']]
            if os.path.isfile("./marksheet/"+input_filename) == False:

                wb = Workbook()
                wb.save("./marksheet/"+input_filename)

            book = load_workbook("./marksheet/"+input_filename)
            for worksheets in book.sheetnames:

                ss_sheet = book[worksheets]

                ss_sheet.title = 'quiz'
                Sheet1 = book['quiz']

                def BOLder(r, c, val, bi, col):
                    if bi:
                        fontStyle = Font(name='Century', size="12",
                                         bold=True, color=col)
                        Sheet1.cell(row=r, column=c,
                                    value=val).font = fontStyle
                    else:
                        fontStyle = Font(name='Century', size="12", color=col)
                        Sheet1.cell(row=r, column=c,
                                    value=val).font = fontStyle

                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                img = openpyxl.drawing.image.Image('image002.png')

                img.anchor = 'A1'

                column = 1
                while column < 10:
                    i = get_column_letter(column)
                    Sheet1.column_dimensions[i].width = 17.8
                    column += 1
                Sheet1.add_image(img)
                Sheet1['A5'].font = Font(bold=True)
                fontStyle = Font(name='Century', size="18", underline='single')
                Sheet1.row_dimensions[5].height = 23
                Sheet1.cell(row=5, column=1,
                            value='Mark Sheet').font = fontStyle
                Sheet1.merge_cells('A5:E5')
                rows = range(1, 44)
                columns = range(1, 10)
                for row in rows:
                    for col in columns:
                        Sheet1.cell(row, col).alignment = Alignment(
                            horizontal='center', vertical='center', wrap_text=True)
                BLACK = 'FF000000'
                BOLder(6, 1, 'Name:', False, BLACK)
                BOLder(6, 2, dct['Name'], True, BLACK)
                Sheet1.merge_cells('B6:C6')
                BOLder(6, 4, 'Exam:', False, BLACK)
                BOLder(6, 5, 'quiz', True, BLACK)
                BOLder(7, 1, 'Roll Number:', False, BLACK)
                BOLder(7, 2, dct['Roll Number'].upper(), True, BLACK)
                Sheet1.merge_cells('B7:C7')

                rows = range(9, 13)
                columns = range(1, 6)
                for row in rows:
                    for col in columns:
                        Sheet1.cell(row=row, column=col).border = thin_border

                BOLder(9, 2, 'Right', True, BLACK)
                BOLder(9, 3, 'Wrong', True, BLACK)
                BOLder(9, 4, 'Not Attempt', True, BLACK)
                BOLder(9, 5, 'Max', True, BLACK)
                BOLder(12, 1, 'Total', True, BLACK)
                BOLder(10, 1, 'No.', True, BLACK)
                BOLder(11, 1, 'Marking', True, BLACK)
                BOLder(15, 1, 'Student Ans', True, BLACK)
                BOLder(15, 2, 'Correct Ans', True, BLACK)
                BOLder(15, 4, 'Student Ans', True, BLACK)
                BOLder(15, 5, 'Correct Ans', True, BLACK)

                student = roll_to_list[dct['Roll Number'].upper()]
                cnt = 0
                na = 0

                for i in range(0, len(ans)):
                    cnt += (student[i] == ans[i])
                    if student[i] == "":
                        na += 1
                RED = 'FFFF0000'
                BLUE = 'FF0000FF'
                DARKGREEN = 'FF008000'
                wng = len(ans) - cnt - na
                BOLder(10, 2, cnt, False, DARKGREEN)
                BOLder(10, 3, wng, False, RED)
                BOLder(10, 4, na, False, BLACK)
                BOLder(10, 5, len(ans), False, BLACK)
                BOLder(11, 4, 0, False, BLACK)
                BOLder(11, 2, pm, False, DARKGREEN)
                BOLder(11, 3, nm, False, RED)
                BOLder(12, 2, pm*cnt, False, DARKGREEN)
                BOLder(12, 3, nm*wng, False, RED)
                BOLder(12, 5, str(pm*cnt + nm*wng) +
                       "/"+str(pm*len(ans)), False, BLUE)

                score = [str(pm*cnt + nm*wng) + "/" +
                         str(pm*len(ans)), cnt, wng, na,str(pm*cnt) + "/" + str(pm*len(ans))]
                list_of_roll_to_score[dct['Roll Number'].upper()] = score

                rows = range(16, 16+len(ans))
                cnt = 0
                for row in rows:
                    Sheet1.cell(row=row, column=1).border = thin_border
                    if(student[cnt] == ans[cnt]):
                        BOLder(row, 1, ans[cnt], False, DARKGREEN)
                    else:
                        BOLder(row, 1, student[cnt], False, RED)
                    Sheet1.cell(row=row, column=2).border = thin_border
                    BOLder(row, 2, ans[cnt], False, BLUE)
                    cnt += 1
                    if(row == 40):
                        break
                row = 16
                for j in range(cnt, len(ans)):
                    Sheet1.cell(row=row, column=4).border = thin_border

                    if(student[j] == ans[j]):
                        BOLder(row, 4, ans[j], False, DARKGREEN)
                    else:
                        BOLder(row, 4, student[j], False, RED)
                    Sheet1.cell(row=row, column=5).border = thin_border
                    BOLder(row, 5, ans[j], False, BLUE)
                    row += 1
                Sheet1.cell(row=15, column=1).border = thin_border
                Sheet1.cell(row=15, column=2).border = thin_border
                Sheet1.cell(row=15, column=4).border = thin_border
                Sheet1.cell(row=15, column=5).border = thin_border
                book.save("./marksheet/"+input_filename)
                break
            # break
    # print(f"{len(list_of_title)} in generate")


def generate2(pm, nm):
    print("HEYYY")
    if 'ANSWER' not in roll_to_list.keys():
        return 0

    with open("master_roll.csv", 'r') as file:
        reader = csv.DictReader(file)

        ans = roll_to_list["ANSWER"]
        for row in reader:
            dct = dict(row)
            if(dct['roll'].upper() not in roll_to_all.keys()):
                print("HEY")
                input_filename = dct['roll'].upper() + ".xlsx"
                if os.path.isfile("./marksheet/"+input_filename) == False:

                    wb = Workbook()
                    wb.save("./marksheet/"+input_filename)

                book = load_workbook("./marksheet/"+input_filename)
                for worksheets in book.sheetnames:

                    ss_sheet = book[worksheets]

                    ss_sheet.title = 'quiz'
                    Sheet1 = book['quiz']

                    def BOLder(r, c, val, bi, col):
                        if bi:
                            fontStyle = Font(name='Century', size="12",
                                             bold=True, color=col)
                            Sheet1.cell(row=r, column=c,
                                        value=val).font = fontStyle
                        else:
                            fontStyle = Font(
                                name='Century', size="12", color=col)
                            Sheet1.cell(row=r, column=c,
                                        value=val).font = fontStyle

                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    img = openpyxl.drawing.image.Image('image002.png')

                    img.anchor = 'A1'

                    column = 1
                    while column < 10:
                        i = get_column_letter(column)
                        Sheet1.column_dimensions[i].width = 17.8
                        column += 1
                    Sheet1.add_image(img)
                    Sheet1['A5'].font = Font(bold=True)
                    fontStyle = Font(name='Century', size="18",
                                     underline='single')
                    Sheet1.row_dimensions[5].height = 23
                    Sheet1.cell(row=5, column=1,
                                value='Mark Sheet').font = fontStyle
                    Sheet1.merge_cells('A5:E5')
                    rows = range(1, 44)
                    columns = range(1, 10)
                    for row in rows:
                        for col in columns:
                            Sheet1.cell(row, col).alignment = Alignment(
                                horizontal='center', vertical='center', wrap_text=True)
                    BLACK = 'FF000000'
                    BOLder(6, 1, 'Name:', False, BLACK)
                    BOLder(6, 2, dct['name'], True, BLACK)
                    Sheet1.merge_cells('B6:C6')
                    BOLder(6, 4, 'Exam:', False, BLACK)
                    BOLder(6, 5, 'quiz', True, BLACK)
                    BOLder(7, 1, 'Roll Number:', False, BLACK)
                    BOLder(7, 2, dct['roll'].upper(), True, BLACK)
                    Sheet1.merge_cells('B7:C7')

                    rows = range(9, 13)
                    columns = range(1, 6)
                    for row in rows:
                        for col in columns:
                            Sheet1.cell(
                                row=row, column=col).border = thin_border

                    BOLder(9, 2, 'Right', True, BLACK)
                    BOLder(9, 3, 'Wrong', True, BLACK)
                    BOLder(9, 4, 'Not Attempt', True, BLACK)
                    BOLder(9, 5, 'Max', True, BLACK)
                    BOLder(12, 1, 'Total', True, BLACK)
                    BOLder(10, 1, 'No.', True, BLACK)
                    BOLder(11, 1, 'Marking', True, BLACK)
                    BOLder(15, 1, 'Student Ans', True, BLACK)
                    BOLder(15, 2, 'Correct Ans', True, BLACK)
                    BOLder(15, 4, 'Student Ans', True, BLACK)
                    BOLder(15, 5, 'Correct Ans', True, BLACK)

                    cnt = 0
                    na = 0
                    wng = 0

                    RED = 'FFFF0000'
                    BLUE = 'FF0000FF'
                    DARKGREEN = 'FF008000'
                    na = len(ans)
                    BOLder(10, 2, cnt, False, DARKGREEN)
                    BOLder(10, 3, wng, False, RED)
                    BOLder(10, 4, na, False, BLACK)
                    BOLder(10, 5, len(ans), False, BLACK)
                    BOLder(11, 4, 0, False, BLACK)
                    BOLder(11, 2, pm, False, DARKGREEN)
                    BOLder(11, 3, nm, False, RED)
                    BOLder(12, 2, pm*cnt, False, DARKGREEN)
                    BOLder(12, 3, nm*wng, False, RED)
                    BOLder(12, 5, str(pm*cnt + nm*wng) +
                           "/"+str(pm*len(ans)), False, BLUE)

                    score = [str(pm*cnt + nm*wng) + "/" +
                             str(pm*len(ans)), cnt, wng, na]
                    list_of_roll_to_score[dct['roll'].upper()] = score

                    rows = range(16, 16+len(ans))
                    cnt = 0
                    for row in rows:
                        Sheet1.cell(row=row, column=1).border = thin_border
                        Sheet1.cell(row=row, column=2).border = thin_border
                        BOLder(row, 2, ans[cnt], False, BLUE)
                        cnt += 1
                        if(row == 40):
                            break
                    row = 16
                    for j in range(cnt, len(ans)):
                        Sheet1.cell(row=row, column=4).border = thin_border
                        Sheet1.cell(row=row, column=5).border = thin_border
                        BOLder(row, 5, ans[j], False, BLUE)
                        row += 1
                    Sheet1.cell(row=15, column=1).border = thin_border
                    Sheet1.cell(row=15, column=2).border = thin_border
                    Sheet1.cell(row=15, column=4).border = thin_border
                    Sheet1.cell(row=15, column=5).border = thin_border
                    book.save("./marksheet/"+input_filename)
                    break


@app.route("/forward", methods=['POST'])
def move_forward():
    pre()
    print("HEyyy")
    if os.path.exists("./marksheet.zip"):
        os.remove("./marksheet.zip")
    if request.form['correct'] == "" or request.form['wrong'] == "":
        flash("Please enter valid input")
        return render_template('index.html')
    pm = float(request.form['correct'])
    nm = float(request.form['wrong'])
    if pm <= 0 or nm > 0:
        flash("Please enter valid input")
        return render_template('index.html')
    print(pm, nm)
    generate(pm, nm)
    if(generate2(pm, nm) == 0):
        flash("No Roll Number with 'ANSWER' is present, Cannot Process!")
        return render_template('index.html', x="")

    flash("Roll wise Generated successfully")
    # print(list_of_roll_to_score)
    # print(f"{len(list_of_title)} in forward")
    print("Roll Wise generated Successfully")

    return render_template('index.html', x="")


@app.route("/upload", methods=['GET', 'POST'])
def upload1():
    if request.method == 'POST':

        try:

            f = request.files['file']
            print(f, f.filename)
            if f.filename == "master_roll.csv":

                if os.path.exists("./master_roll.csv"):
                    os.remove("./master_roll.csv")
            if f.filename == "responses.csv":

                if os.path.exists("./responses.csv"):
                    os.remove("./responses.csv")

            f.save(f.filename)
        except:
            flash("You have not selected any file 😲")
            return render_template('index.html')
        flash(f"{f.filename} uploaded successfully 👍")

    return render_template('index.html', x="disabled")


@app.route("/", methods=['GET', 'POST'])
def hello_world():
    roll_to_email.clear()
    roll_to_all.clear()
    list_of_roll_to_score.clear()
    roll_to_list.clear()
    list_of_roll_to_score.clear()
    list_of_title.clear()
    return render_template('index.html', x="disabled")


if __name__ == "__main__":
    app.run(debug=True, port=8000)
